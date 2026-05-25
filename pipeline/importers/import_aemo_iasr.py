"""
Import AEMO Inputs and Assumptions Workbook (IASR Workbook).

The IASR Workbook is AEMO's master spreadsheet of planning assumptions used
for each Integrated System Plan (ISP). The "Existing Gen Data Summary" sheet
includes per-project rows for every generator in the NEM model:

  * Existing (459)                           — already operating
  * Committed (90)                           — committed to construction
  * Anticipated (56)                         — likely to commit
  * Additional policy-supported (38)         — included in policy scenarios

Critical added value vs the monthly AEMO Gen Information Excel: the IASR
exposes **REZ Location** + **REZ ID** per project (e.g., 'Wagga Wagga' /
'N5', 'Central-West Orana' / 'N3') — the monthly Gen Info Excel does NOT.

Source: https://www.aemo.com.au/-/media/files/stakeholder_consultation/consultations/nem-consultations/2024/2025-iasr-scenarios/final-docs/2025-inputs-and-assumptions-workbook.xlsx
Cadence: Annual (final IASR — August), with Draft IASR cycles in Dec/Feb
         supporting each ISP. The 2025 IASR Workbook supports the 2026 ISP.

Pipeline-data-protection rule: the importer uses COALESCE to fill `projects.rez`
only when currently NULL/empty — never overwriting a hand-curated value.
"""
import os
import re
import sys
from urllib.request import urlopen, Request

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from db import get_connection, init_db


WORKBOOK_URL = (
    'https://www.aemo.com.au/-/media/files/stakeholder_consultation/'
    'consultations/nem-consultations/2024/2025-iasr-scenarios/final-docs/'
    '2025-inputs-and-assumptions-workbook.xlsx'
)
LOCAL_FILE = os.path.join(os.path.dirname(__file__), 'downloads', 'aemo-2025-iasr-workbook.xlsx')
WORKBOOK_VERSION = '2025-iasr-aug2025'

# We are interested in pipeline projects (not the existing 459) — these are the
# rows where IASR adds value via REZ enrichment.
PIPELINE_STATUSES = {'Committed', 'Anticipated', 'Additional policy-supported'}

# Column positions (1-indexed) in 'Existing Gen Data Summary' sheet, header at row 10
COL = {
    'iasr_id':         2,
    'power_station':   3,
    'tech_type':       4,
    'fuel_type':       5,
    'region':          6,
    'sub_region':      7,
    'rez_location':    8,
    'rez_id':          9,
    'status':         10,
    'max_capacity_mw': 12,
    'storage_mwh':    13,
}
DATA_START_ROW = 13


# Tech-type normalization for matching (IASR uses verbose labels)
def normalise_tech(iasr_tech):
    if not iasr_tech:
        return None
    t = iasr_tech.lower()
    if 'solar' in t:           return 'solar'
    if 'wind' in t:            return 'wind'
    if 'battery' in t:         return 'bess'
    if 'pumped hydro' in t:    return 'pumped_hydro'
    if 'ocgt' in t or 'ccgt' in t or 'gas' in t: return 'gas'
    return None


def download_if_missing():
    if os.path.exists(LOCAL_FILE):
        size_mb = os.path.getsize(LOCAL_FILE) / 1024 / 1024
        print(f"  File already exists: {LOCAL_FILE} ({size_mb:.1f} MB)")
        return
    print(f"  Downloading from {WORKBOOK_URL} ...")
    os.makedirs(os.path.dirname(LOCAL_FILE), exist_ok=True)
    # AEMO blocks the default urllib UA — pass a browser-like UA
    req = Request(WORKBOOK_URL, headers={
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/605.1.15 Safari/605.1.15'
    })
    with urlopen(req) as resp, open(LOCAL_FILE, 'wb') as out:
        out.write(resp.read())
    size_mb = os.path.getsize(LOCAL_FILE) / 1024 / 1024
    print(f"  Downloaded {size_mb:.1f} MB")


def parse_workbook(path):
    """Yield dicts for each pipeline (Committed/Anticipated/Policy-supported) row."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb['Existing Gen Data Summary']
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row, values_only=True):
        # row is a tuple; col indices are 1-based, so subtract 1 to index
        status = row[COL['status'] - 1] if len(row) > COL['status'] - 1 else None
        if status not in PIPELINE_STATUSES:
            continue
        iasr_id = row[COL['iasr_id'] - 1]
        power_station = row[COL['power_station'] - 1]
        if not iasr_id or not power_station:
            continue
        rez_location = row[COL['rez_location'] - 1]
        rez_id = row[COL['rez_id'] - 1]
        # 'Not Applicable' → None for REZ fields
        if rez_location in ('Not Applicable', None, ''): rez_location = None
        if rez_id in ('Not Applicable', None, ''):       rez_id = None
        yield {
            'iasr_id':         str(iasr_id).strip(),
            'power_station':   str(power_station).strip(),
            'tech_type':       row[COL['tech_type'] - 1],
            'fuel_type':       row[COL['fuel_type'] - 1],
            'region':          row[COL['region'] - 1],
            'sub_region':      row[COL['sub_region'] - 1],
            'rez_location':    rez_location,
            'rez_id':          rez_id,
            'status':          status,
            'max_capacity_mw': row[COL['max_capacity_mw'] - 1],
            'storage_mwh':     row[COL['storage_mwh'] - 1],
        }


def slug(s):
    """Normalise a string for fuzzy matching."""
    s = (s or '').lower()
    s = re.sub(r'\s*\(.*?\)', '', s)  # strip parentheticals e.g. ' (Stage 2)'
    s = re.sub(r'[^a-z0-9 ]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    # Common suffix strips
    for suffix in [' farm', ' power station', ' wind', ' solar', ' bess', ' battery']:
        if s.endswith(suffix):
            s = s[:-len(suffix)]
    return s


def build_project_index(cur):
    """Build a name → project_id lookup for fuzzy matching."""
    cur.execute("""
        SELECT id, name, technology, capacity_mw, state, aemo_gen_info_id
          FROM projects
    """)
    rows = cur.fetchall()
    # By exact name
    by_name = {r['name'].lower(): r['id'] for r in rows}
    # By slugified name + tech
    by_slug = {}
    for r in rows:
        key = (slug(r['name']), r['technology'])
        by_slug.setdefault(key, []).append(r['id'])
    # By AEMO Gen Info ID (DUID or station)
    by_aemo = {str(r['aemo_gen_info_id']).strip(): r['id'] for r in rows if r['aemo_gen_info_id']}
    return by_name, by_slug, by_aemo


def match_project(row, by_name, by_slug, by_aemo):
    """Try exact name → slug+tech → AEMO ID."""
    ps_lower = row['power_station'].lower()
    if ps_lower in by_name:
        return by_name[ps_lower]
    tech = normalise_tech(row['tech_type'])
    if tech:
        key = (slug(row['power_station']), tech)
        if key in by_slug and len(by_slug[key]) == 1:
            return by_slug[key][0]
    # Fall back: AEMO IASR ID may match a project's aemo_gen_info_id
    if row['iasr_id'] in by_aemo:
        return by_aemo[row['iasr_id']]
    return None


def upsert_iasr_project(cur, row, project_id):
    cur.execute("""
        INSERT INTO aemo_iasr_projects (
            iasr_id, project_id, power_station, technology_type, fuel_type,
            region, sub_region, rez_location, rez_id, status,
            max_capacity_mw, storage_mwh, workbook_version, last_imported_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
        ON CONFLICT(iasr_id, workbook_version) DO UPDATE SET
            project_id        = excluded.project_id,
            power_station     = excluded.power_station,
            technology_type   = excluded.technology_type,
            fuel_type         = excluded.fuel_type,
            region            = excluded.region,
            sub_region        = excluded.sub_region,
            rez_location      = excluded.rez_location,
            rez_id            = excluded.rez_id,
            status            = excluded.status,
            max_capacity_mw   = excluded.max_capacity_mw,
            storage_mwh       = excluded.storage_mwh,
            last_imported_at  = datetime('now')
    """, (
        row['iasr_id'], project_id, row['power_station'], row['tech_type'], row['fuel_type'],
        row['region'], row['sub_region'], row['rez_location'], row['rez_id'], row['status'],
        row['max_capacity_mw'], row['storage_mwh'], WORKBOOK_VERSION,
    ))


def backfill_project_rez(cur):
    """Update projects.rez from IASR rez_location where projects.rez is NULL.

    COALESCE protection: never overwrite a hand-curated value.
    """
    cur.execute("""
        UPDATE projects
           SET rez = (
               SELECT rez_location FROM aemo_iasr_projects
                WHERE project_id = projects.id
                  AND rez_location IS NOT NULL
                ORDER BY last_imported_at DESC LIMIT 1
           )
         WHERE (rez IS NULL OR rez = '')
           AND EXISTS (
               SELECT 1 FROM aemo_iasr_projects
                WHERE project_id = projects.id
                  AND rez_location IS NOT NULL
           )
    """)
    return cur.rowcount


def main():
    print("Importing AEMO IASR Workbook (Existing Gen Data Summary)")
    print("=" * 60)

    download_if_missing()
    init_db()
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO import_runs (source, source_file, status)
        VALUES ('aemo_iasr', ?, 'running')
    """, (LOCAL_FILE,))
    run_id = cur.lastrowid

    print("\n  Parsing workbook ...")
    rows = list(parse_workbook(LOCAL_FILE))
    print(f"  Found {len(rows)} pipeline projects (Committed/Anticipated/Policy-supported)")

    by_name, by_slug, by_aemo = build_project_index(cur)

    matched   = 0
    unmatched = []
    for row in rows:
        project_id = match_project(row, by_name, by_slug, by_aemo)
        if project_id:
            matched += 1
        else:
            unmatched.append(row)
        upsert_iasr_project(cur, row, project_id)

    rez_backfilled = backfill_project_rez(cur)

    cur.execute("""
        UPDATE import_runs
           SET records_imported = ?, records_updated = ?,
               status = 'completed', completed_at = datetime('now')
         WHERE id = ?
    """, (len(rows), matched, run_id))

    conn.commit()

    print(f"\n  Import complete:")
    print(f"    IASR rows ingested:     {len(rows)}")
    print(f"    Matched to projects:    {matched}")
    print(f"    Unmatched:              {len(unmatched)}")
    print(f"    Projects rez backfilled:{rez_backfilled}")

    if unmatched:
        print(f"\n  ⚠ Top 15 unmatched (may need PROJECT_MAP entries):")
        for r in unmatched[:15]:
            print(f"    {r['iasr_id']:14s} | {r['power_station']:36s} | {r['status']:13s} | "
                  f"{(r['rez_location'] or '-'):22s} | {r['region']}")

    cur.execute("""
        SELECT status, COUNT(*) AS n FROM aemo_iasr_projects
         WHERE workbook_version = ? GROUP BY status ORDER BY n DESC
    """, (WORKBOOK_VERSION,))
    print(f"\n  Status distribution (this workbook):")
    for r in cur.fetchall():
        print(f"    {r['status']:35s} : {r['n']}")

    cur.execute("""
        SELECT rez_location, COUNT(*) AS n, SUM(max_capacity_mw) AS mw
          FROM aemo_iasr_projects
         WHERE workbook_version = ?
           AND project_id IS NOT NULL
           AND rez_location IS NOT NULL
         GROUP BY rez_location
         ORDER BY mw DESC
         LIMIT 10
    """, (WORKBOOK_VERSION,))
    print(f"\n  Top 10 REZ locations (matched projects, by MW):")
    for r in cur.fetchall():
        print(f"    {r['rez_location']:30s} | {r['n']:>3d} projects | {r['mw']:>8.1f} MW")

    conn.close()
    print("\n" + "=" * 60)
    print("Done!")


if __name__ == '__main__':
    main()
