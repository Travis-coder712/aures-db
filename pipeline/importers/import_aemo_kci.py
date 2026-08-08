#!/usr/bin/env python3
"""
AEMO Key Connection Information (KCI) importer.

The KCI dataset is the per-project connection-application registry that every
TNSP submits to AEMO quarterly under NER 3.7F(3). AEMO publishes both per-TNSP
Excel files and an all-NEM compiled workbook on the Generation Information
page. This importer ingests the compiled workbook.

The KCI carries data no other AEMO source does:
  - NSP (TransGrid / Powerlink / VicGrid / ElectraNet / TasNetworks) per project
  - AEMO KCI ID — the stable join key that also appears in Rosetta and future
    KCI releases
  - Application status (Active / Complete / Withdrawn) + application type
  - Forecast COD (earliest + latest)
  - Energy conversion technology subtype (single-axis tracking, grid-forming
    inverter, etc.)
  - Per-generating-unit MW and count (turbine count for wind)
  - Sync / async classification

The file DOES NOT contain a discrete Connection Point / Substation column.
Substation data is in the free-text "Site Location Description" for some rows
but not consistently structured. Substation-level detail still comes from
the per-TNSP TAPR PDFs (see docs/RESEARCH_EIS_COVERAGE_AUDIT.md).

Idempotency: UNIQUE(snapshot, aemo_kci_id, application_id). Same file
re-imported = zero net delta.

Usage:
    python3 pipeline/importers/import_aemo_kci.py
    python3 pipeline/importers/import_aemo_kci.py --snapshot 2026-Q2
"""

import argparse
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from db import get_connection, DB_PATH

try:
    import openpyxl
except ImportError:
    print('ERROR: openpyxl not installed. Run: pip install openpyxl')
    sys.exit(1)

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
KCI_ARCHIVE = os.path.join(REPO_ROOT, 'data', 'gi_snapshots')

# ────────────────────────────────────────────────────────────────────────────
# Column indices in the compiled KCI workbook. Header row = 3, data starts 4.
# ────────────────────────────────────────────────────────────────────────────

COL = {
    'tnsp_name': 0,
    'compilation_timestamp': 1,
    'application_id': 2,
    'application_type': 3,
    'kci_notification_date': 4,
    'kci_validation_date': 5,
    'application_status': 6,
    'organisation_name': 7,
    'abn': 8,
    'acn': 9,
    'site_name': 10,
    'site_location_description': 11,
    'region': 12,
    'max_gen_mw_lower': 13,
    'max_gen_mw_upper': 14,
    'forecast_cod_earliest': 15,
    'forecast_cod_latest': 16,
    'energy_conv_tech': 17,
    'energy_conv_subtype': 18,
    'n_units_lower': 19,
    'n_units_upper': 20,
    'electricity_gen_tech': 21,
    'per_unit_mw_lower': 22,
    'per_unit_mw_upper': 23,
    'nameplate_mw_lower': 24,
    'nameplate_mw_upper': 25,
    'aemo_kci_id': 26,
}

# ────────────────────────────────────────────────────────────────────────────
# Normalisation tables — the KCI data has variant spellings that must be
# collapsed before joining downstream.
# ────────────────────────────────────────────────────────────────────────────

TNSP_NORMAL = {
    'transgrid': 'TransGrid',
    'powerlink': 'Powerlink',
    'vicgrid': 'VicGrid',
    'ausnet': 'AusNet',              # older KCI files used AusNet — treat as VicGrid or preserve
    'ausnet services': 'AusNet',
    'electranet': 'ElectraNet',
    'tasnetworks': 'TasNetworks',
    'tas networks': 'TasNetworks',
}

TECH_NORMAL = {
    'solar pv':         'Solar PV',
    'solarpv':          'Solar PV',
    'solar':            'Solar PV',
    'solar farm':       'Solar PV',
    'solar thermal':    'Solar Thermal',
    'pv + bess':        'Solar + Storage',
    'solar + storage':  'Solar + Storage',
    'battery; solar farm': 'Solar + Storage',
    'wind':             'Wind',
    'wind turbine':     'Wind',
    'windturbine':      'Wind',
    'wind farm':        'Wind',
    'wind farm':        'Wind',
    'battery; wind farm': 'Wind + Storage',
    'storage':          'Storage',
    'battery':          'Storage',
    'hydro':            'Hydro',
    'gas':              'Gas',
    'reciprocating engine': 'Reciprocating Engine',
    'steam turbine':    'Steam Turbine',
    'turbine':          'Gas',       # standalone Turbine tokens in the file are always OCGT/CCGT
    'hydrogen':         'Hydrogen',
    'inverter based resource': 'Inverter Based Resource',
    'grid forming inverter':   'Grid Forming Inverter',
    'other':            'Other',
    'others':           'Other',
    'n/a':              None,
}

NULLISH = {'', 'not provided', 'not Provided', 'Not provided', 'Not Provided',
           'n/a', 'N/A', 'na', 'NA', 'nil', 'None', None, 'NP', 'np'}


# ────────────────────────────────────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────────────────────────────────────

def norm_tnsp(v):
    if v is None: return None
    return TNSP_NORMAL.get(str(v).strip().lower(), str(v).strip())


def norm_tech(v):
    if v is None: return None
    return TECH_NORMAL.get(str(v).strip().lower(), str(v).strip())


def norm_text(v):
    """Collapse various 'Not Provided' variants to None; strip whitespace."""
    if v is None: return None
    s = str(v).strip()
    if s in NULLISH: return None
    return s


def norm_number(v):
    """Return float if numeric, else None. Handles strings like '100', '3.437', 'n/a'."""
    if v is None: return None
    if isinstance(v, (int, float)):
        return float(v) if v != 0 or v == 0 else None
    s = str(v).strip()
    if s in NULLISH: return None
    try:
        return float(s.replace(',', ''))
    except ValueError:
        return None


def norm_int(v):
    n = norm_number(v)
    return int(n) if n is not None else None


def norm_date(v):
    if v is None: return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if s in NULLISH: return None
    # Try 'YYYY-MM-DD HH:MM:SS' or similar
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return s  # store as-is if unrecognised


def slugify(name):
    """Match the AURES slugify convention."""
    if not name: return ''
    s = str(name).lower().strip()
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()
    s = re.sub(r'[^a-z0-9\s-]', '', s)
    s = re.sub(r'[\s-]+', '-', s).strip('-')
    return s


# ────────────────────────────────────────────────────────────────────────────
# Schema migration
# ────────────────────────────────────────────────────────────────────────────

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS aemo_kci_records (
    id                          INTEGER PRIMARY KEY AUTOINCREMENT,
    snapshot                    TEXT NOT NULL,               -- e.g. '2026-Q2'
    aemo_kci_id                 TEXT,                        -- N00001 etc — the KCI join key
    project_id                  TEXT REFERENCES projects(id),-- matched via Site Name — nullable

    tnsp_name                   TEXT,                        -- normalised: TransGrid / Powerlink / VicGrid / ElectraNet / TasNetworks
    application_id              TEXT,                        -- TNSP application ID
    application_status          TEXT,                        -- Active / Complete
    application_type            TEXT,                        -- 'New application to connect (NER 5.3)' etc

    organisation_name           TEXT,
    abn                         TEXT,
    site_name                   TEXT,
    site_location_description   TEXT,
    region                      TEXT,                        -- NSW1 / QLD1 / VIC1 / SA1 / TAS1

    max_gen_mw_lower            REAL,
    max_gen_mw_upper            REAL,
    forecast_cod_earliest       TEXT,
    forecast_cod_latest         TEXT,

    energy_conv_tech            TEXT,                        -- normalised
    energy_conv_subtype         TEXT,
    n_units_lower               INTEGER,
    n_units_upper               INTEGER,
    electricity_gen_tech        TEXT,                        -- Asynchronous / Synchronous
    per_unit_mw_lower           REAL,
    per_unit_mw_upper           REAL,
    nameplate_mw_lower          REAL,
    nameplate_mw_upper          REAL,

    kci_notification_date       TEXT,
    kci_validation_date         TEXT,
    compilation_timestamp       TEXT,

    created_at                  TEXT NOT NULL DEFAULT (datetime('now'))
);

CREATE UNIQUE INDEX IF NOT EXISTS idx_aemo_kci_natural_key
  ON aemo_kci_records(snapshot, COALESCE(aemo_kci_id,''), COALESCE(application_id,''));
CREATE INDEX IF NOT EXISTS idx_aemo_kci_project  ON aemo_kci_records(project_id);
CREATE INDEX IF NOT EXISTS idx_aemo_kci_kci_id   ON aemo_kci_records(aemo_kci_id);
CREATE INDEX IF NOT EXISTS idx_aemo_kci_site     ON aemo_kci_records(site_name);
"""


def add_project_column_if_missing(conn):
    """Add projects.aemo_kci_id column if it doesn't already exist."""
    cols = [r[1] for r in conn.execute('PRAGMA table_info(projects)')]
    if 'aemo_kci_id' not in cols:
        conn.execute('ALTER TABLE projects ADD COLUMN aemo_kci_id TEXT')


# ────────────────────────────────────────────────────────────────────────────
# Project matching — Site Name → projects.id
# ────────────────────────────────────────────────────────────────────────────

REGION_TO_STATE = {'NSW1': 'NSW', 'QLD1': 'QLD', 'VIC1': 'VIC', 'SA1': 'SA', 'TAS1': 'TAS'}

# Words that carry no discriminating information — drop before token matching
STOPWORDS = {
    'farm', 'project', 'stage', 'kci', 'energy', 'hub', 'park', 'power',
    'station', 'and', 'the', 'of', 'pty', 'ltd', 'plant', 'facility',
    'renewable', 'renewables', 'generation', 'grid',
    # Tech-marker tokens (present in both KCI and AURES names — noise for matching)
    'solar', 'wind', 'bess', 'battery', 'hybrid', 'storage', 'pumped', 'hydro',
    'pv', 'gas',
    # Ordinals / brand suffixes
    '1', '2', '3', 'i', 'ii', 'north', 'south', 'east', 'west',
}

# Additional aliases to nudge naming variance
ALIASES = [
    (r'pumped[- ]hydro[- ]energy[- ]storage', 'pumped-storage'),
    (r'grid[- ]battery',                      'bess'),
    (r'energy[- ]storage',                    'bess'),
    (r'battery[- ]energy[- ]storage[- ]system','bess'),
]


def canonical_tokens(name):
    """Lowercase; drop stopwords; return non-noise tokens for overlap match."""
    if not name: return set()
    s = str(name).lower()
    for pattern, repl in ALIASES:
        s = re.sub(pattern, repl, s)
    # Split on any non-alphanum
    toks = re.findall(r'[a-z0-9]+', s)
    return {t for t in toks if t not in STOPWORDS and len(t) >= 3}


def build_project_index(conn):
    """Index projects by normalised slug + a token-overlap index keyed by state."""
    by_norm_name = {}
    # By state, list of (project_id, project_name, project_tokens)
    by_state = defaultdict(list)
    for row in conn.execute('SELECT id, name, state, technology FROM projects'):
        pid, name, state, tech = row['id'], row['name'], row['state'], row['technology']
        norm = slugify(name)
        if norm:
            by_norm_name.setdefault(norm, pid)
        tokens = canonical_tokens(name)
        if state:
            by_state[state].append((pid, name, tech, tokens))
    return by_norm_name, by_state


def try_match(site_name, region, by_norm_name, by_state):
    """Return (project_id, match_quality) for a KCI Site Name, or (None, None)."""
    if not site_name:
        return None, None
    site_slug = slugify(site_name)

    # 1. Exact slug match
    pid = by_norm_name.get(site_slug)
    if pid: return pid, 'exact'

    # 2. Try with common suffixes appended
    for suffix in ('-solar-farm', '-wind-farm', '-bess', '-battery', '-solar', '-wind',
                   '-hybrid', '-pumped-storage'):
        if not site_slug.endswith(suffix):
            pid = by_norm_name.get(site_slug + suffix)
            if pid: return pid, 'suffix-added'

    # 3. Strip common KCI-only suffixes
    stripped = re.sub(r'-(kci|project|energy-park|energy-hub|hybrid|stage-\d+)+$', '', site_slug)
    if stripped != site_slug:
        pid = by_norm_name.get(stripped)
        if pid: return pid, 'suffix-stripped'

    # 4. Token-overlap fuzzy match, constrained to the same state
    state = REGION_TO_STATE.get(region)
    if not state:
        return None, None
    site_tokens = canonical_tokens(site_name)
    if not site_tokens:
        return None, None
    candidates = by_state.get(state, [])
    best_pid, best_score = None, 0
    for pid, name, tech, ptokens in candidates:
        if not ptokens: continue
        overlap = len(site_tokens & ptokens)
        if overlap == 0: continue
        # Score = overlap / min(sizes) — 1.0 = one set fully contains the other
        score = overlap / min(len(site_tokens), len(ptokens))
        if score > best_score:
            best_score, best_pid = score, pid
    # Require perfect containment (score == 1.0) with at least 1 token to match
    if best_score >= 1.0:
        return best_pid, 'token-overlap'
    return None, None


# ────────────────────────────────────────────────────────────────────────────
# Main importer
# ────────────────────────────────────────────────────────────────────────────

def import_kci(filepath, snapshot):
    conn = get_connection()

    print(f'\n[1/5] Migrating schema...')
    conn.executescript(SCHEMA_SQL)
    add_project_column_if_missing(conn)
    conn.commit()

    print(f'[2/5] Parsing {os.path.basename(filepath)}...')
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]  # single sheet

    rows_seen = 0
    rows_imported = 0
    by_norm_name, by_state = build_project_index(conn)
    matched_project_ids = set()
    unmatched_examples = []
    tnsp_counts = Counter()
    tech_counts = Counter()
    match_status = Counter()
    match_quality = Counter()

    for raw in ws.iter_rows(min_row=4, values_only=True):
        if not any(raw):
            continue
        rows_seen += 1

        site_name = norm_text(raw[COL['site_name']])
        region = norm_text(raw[COL['region']])
        pid, quality = try_match(site_name, region, by_norm_name, by_state)
        if pid:
            matched_project_ids.add(pid)
            match_status['matched'] += 1
            match_quality[quality] += 1
        else:
            match_status['unmatched'] += 1
            if site_name and len(unmatched_examples) < 20:
                unmatched_examples.append(site_name)

        rec = {
            'snapshot': snapshot,
            'aemo_kci_id': norm_text(raw[COL['aemo_kci_id']]),
            'project_id': pid,
            'tnsp_name': norm_tnsp(raw[COL['tnsp_name']]),
            'application_id': norm_text(raw[COL['application_id']]),
            'application_status': norm_text(raw[COL['application_status']]),
            'application_type': norm_text(raw[COL['application_type']]),
            'organisation_name': norm_text(raw[COL['organisation_name']]),
            'abn': norm_text(raw[COL['abn']]),
            'site_name': site_name,
            'site_location_description': norm_text(raw[COL['site_location_description']]),
            'region': norm_text(raw[COL['region']]),
            'max_gen_mw_lower': norm_number(raw[COL['max_gen_mw_lower']]),
            'max_gen_mw_upper': norm_number(raw[COL['max_gen_mw_upper']]),
            'forecast_cod_earliest': norm_date(raw[COL['forecast_cod_earliest']]),
            'forecast_cod_latest': norm_date(raw[COL['forecast_cod_latest']]),
            'energy_conv_tech': norm_tech(raw[COL['energy_conv_tech']]),
            'energy_conv_subtype': norm_text(raw[COL['energy_conv_subtype']]),
            'n_units_lower': norm_int(raw[COL['n_units_lower']]),
            'n_units_upper': norm_int(raw[COL['n_units_upper']]),
            'electricity_gen_tech': norm_text(raw[COL['electricity_gen_tech']]),
            'per_unit_mw_lower': norm_number(raw[COL['per_unit_mw_lower']]),
            'per_unit_mw_upper': norm_number(raw[COL['per_unit_mw_upper']]),
            'nameplate_mw_lower': norm_number(raw[COL['nameplate_mw_lower']]),
            'nameplate_mw_upper': norm_number(raw[COL['nameplate_mw_upper']]),
            'kci_notification_date': norm_date(raw[COL['kci_notification_date']]),
            'kci_validation_date': norm_date(raw[COL['kci_validation_date']]),
            'compilation_timestamp': str(raw[COL['compilation_timestamp']]) if raw[COL['compilation_timestamp']] else None,
        }

        if rec['tnsp_name']: tnsp_counts[rec['tnsp_name']] += 1
        if rec['energy_conv_tech']: tech_counts[rec['energy_conv_tech']] += 1

        cols = list(rec.keys())
        vals = list(rec.values())
        placeholders = ','.join(['?'] * len(vals))
        update_clause = ', '.join(f'{c} = excluded.{c}' for c in cols
                                  if c not in ('snapshot', 'aemo_kci_id', 'application_id'))
        conn.execute(
            f'''INSERT INTO aemo_kci_records ({",".join(cols)}) VALUES ({placeholders})
                ON CONFLICT(snapshot, COALESCE(aemo_kci_id,''), COALESCE(application_id,''))
                DO UPDATE SET {update_clause}''',
            vals,
        )
        rows_imported += 1

    print(f'[3/5] Ingested {rows_imported} rows from {rows_seen} data rows in workbook')
    print(f'      TNSPs: {dict(tnsp_counts)}')
    print(f'      Tech:  {dict(tech_counts)}')
    print(f'      Match: {dict(match_status)} ({len(matched_project_ids)} distinct AURES projects matched)')
    print(f'      Match quality: {dict(match_quality)}')

    # Populate projects.aemo_kci_id from the most recent Active record per project
    print(f'\n[4/5] Populating projects.aemo_kci_id + connection_nsp from KCI (Active records preferred)...')
    conn.execute('''
      UPDATE projects
         SET aemo_kci_id = (
               SELECT aemo_kci_id FROM aemo_kci_records
                WHERE project_id = projects.id
                  AND aemo_kci_id IS NOT NULL
                ORDER BY CASE application_status WHEN 'Active' THEN 0 ELSE 1 END,
                         kci_validation_date DESC
                LIMIT 1
             )
       WHERE id IN (SELECT DISTINCT project_id FROM aemo_kci_records WHERE project_id IS NOT NULL)
    ''')
    conn.execute('''
      UPDATE projects
         SET connection_nsp = (
               SELECT tnsp_name FROM aemo_kci_records
                WHERE project_id = projects.id
                  AND tnsp_name IS NOT NULL
                ORDER BY CASE application_status WHEN 'Active' THEN 0 ELSE 1 END,
                         kci_validation_date DESC
                LIMIT 1
             )
       WHERE (connection_nsp IS NULL OR connection_nsp = '')
         AND id IN (SELECT DISTINCT project_id FROM aemo_kci_records WHERE project_id IS NOT NULL)
    ''')
    n_kci_pop = conn.execute('SELECT COUNT(*) FROM projects WHERE aemo_kci_id IS NOT NULL').fetchone()[0]
    n_nsp_pop = conn.execute("SELECT COUNT(*) FROM projects WHERE connection_nsp IS NOT NULL AND connection_nsp != ''").fetchone()[0]
    print(f'      projects.aemo_kci_id populated:    {n_kci_pop}')
    print(f'      projects.connection_nsp populated: {n_nsp_pop}')

    conn.commit()

    # Match-quality report on unmatched
    print(f'\n[5/5] Unmatched Site Names (first 20 of {match_status["unmatched"]}):')
    for s in unmatched_examples:
        print(f'      - {s}')

    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--snapshot', default='2026-Q2',
                        help="KCI release label like '2026-Q2' (default: 2026-Q2)")
    parser.add_argument('--file', default=None,
                        help='Explicit path to KCI xlsx. Defaults to data/gi_snapshots/kci_<snapshot>/kci-nem-compiled-<snapshot>.xlsx')
    args = parser.parse_args()

    # Path uses the tighter form (kci_2026Q2/kci-nem-compiled-2026Q2.xlsx);
    # the snapshot label uses the hyphenated form (2026-Q2) for consistency
    # with the AEMO GI snapshot convention.
    path_tag = args.snapshot.replace('-', '')
    fname = f'kci-nem-compiled-{path_tag}.xlsx'
    filepath = args.file or os.path.join(KCI_ARCHIVE, f'kci_{path_tag}', fname)
    if not os.path.exists(filepath):
        print(f'ERROR: KCI file not found at {filepath}')
        sys.exit(1)

    print('=' * 60)
    print('AEMO KCI Importer')
    print('=' * 60)
    print(f'  Snapshot: {args.snapshot}')
    print(f'  Source:   {filepath}')

    import_kci(filepath, args.snapshot)

    print(f'\n{"=" * 60}\nDone.\n{"=" * 60}')


if __name__ == '__main__':
    main()
