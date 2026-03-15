#!/usr/bin/env python3
"""
AEMO Integrated System Plan (ISP) Data Importer
================================================
Reads ISP appendix Excel files for REZ hosting capacity and connection data.

The ISP files should be placed in pipeline/importers/downloads/
Expected files: ISP appendix workbooks (xlsx format)

Usage:
    python3 pipeline/importers/import_aemo_isp.py
    python3 pipeline/importers/import_aemo_isp.py --file path/to/isp.xlsx
"""

import os
import sys
import argparse
from datetime import datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from db import get_connection

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

DOWNLOADS_DIR = os.path.join(os.path.dirname(__file__), 'downloads')

# REZ ID mappings from ISP names to our slug format
REZ_NAME_MAP = {
    'Central-West Orana': 'nsw-central-west-orana',
    'CWO': 'nsw-central-west-orana',
    'New England': 'nsw-new-england',
    'South West NSW': 'nsw-south-west',
    'SW NSW': 'nsw-south-west',
    'Hunter-Central Coast': 'nsw-hunter-central-coast',
    'Illawarra': 'nsw-illawarra',
    'North Queensland': 'qld-north-queensland',
    'North QLD': 'qld-north-queensland',
    'Isaac': 'qld-isaac',
    'Fitzroy': 'qld-fitzroy',
    'Darling Downs': 'qld-darling-downs',
    'Wide Bay': 'qld-wide-bay',
    'South East Queensland': 'qld-south-east',
    'SE QLD': 'qld-south-east',
    'Gippsland': 'vic-gippsland',
    'Murray River': 'vic-murray-river',
    'Western Victoria': 'vic-western-victoria',
    'South West Victoria': 'vic-south-west-victoria',
    'SW VIC': 'vic-south-west-victoria',
    'Mid-North SA': 'sa-mid-north',
    'South East SA': 'sa-south-east',
    'Leigh Creek': 'sa-leigh-creek',
    'Roxby Downs': 'sa-roxby-downs',
    'North West Tasmania': 'tas-north-west',
    'NW TAS': 'tas-north-west',
    'North East Tasmania': 'tas-north-east',
    'NE TAS': 'tas-north-east',
}


def ensure_isp_table(conn):
    """Create the rez_isp_data table if it doesn't exist."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS rez_isp_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rez_id TEXT NOT NULL,
            isp_year INTEGER NOT NULL,
            hosting_capacity_mw REAL,
            connection_capacity_mw REAL,
            transmission_status TEXT,
            expected_available TEXT,
            notes TEXT,
            import_date TEXT DEFAULT (date('now')),
            UNIQUE(rez_id, isp_year)
        )
    """)
    conn.commit()


def normalize_rez_name(name):
    """Try to match an ISP REZ name to our slug format."""
    if not name:
        return None
    name = str(name).strip()
    # Direct match
    if name in REZ_NAME_MAP:
        return REZ_NAME_MAP[name]
    # Case-insensitive match
    for k, v in REZ_NAME_MAP.items():
        if k.lower() == name.lower():
            return v
    # Partial match
    for k, v in REZ_NAME_MAP.items():
        if k.lower() in name.lower() or name.lower() in k.lower():
            return v
    return None


def import_isp_excel(conn, filepath, isp_year=2024):
    """Parse an ISP Excel workbook and extract REZ data."""
    print(f"\nReading ISP file: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, data_only=True)

    imported = 0
    skipped = 0

    # Look for sheets containing REZ data
    rez_sheets = []
    for sheet_name in wb.sheetnames:
        sl = sheet_name.lower()
        if any(kw in sl for kw in ['rez', 'renewable energy zone', 'hosting', 'connection']):
            rez_sheets.append(sheet_name)

    if not rez_sheets:
        # Try first sheet as fallback
        rez_sheets = [wb.sheetnames[0]]

    for sheet_name in rez_sheets:
        ws = wb[sheet_name]
        print(f"  Processing sheet: {sheet_name}")

        # Find header row — look for "REZ" or "Zone" in first 10 rows
        header_row = None
        header_map = {}
        for row_idx in range(1, min(ws.max_row + 1, 15)):
            for col_idx in range(1, min(ws.max_column + 1, 20)):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = str(cell.value).lower() if cell.value else ''
                if 'rez' in val or 'zone' in val or 'region' in val:
                    header_row = row_idx
                    break
            if header_row:
                break

        if not header_row:
            print(f"    Could not find header row in {sheet_name}")
            continue

        # Map column headers
        for col_idx in range(1, ws.max_column + 1):
            val = str(ws.cell(row=header_row, column=col_idx).value or '').lower()
            if 'rez' in val or 'zone' in val or 'region' in val:
                header_map['name'] = col_idx
            elif 'hosting' in val and 'capacity' in val:
                header_map['hosting_capacity'] = col_idx
            elif 'connection' in val and 'capacity' in val:
                header_map['connection_capacity'] = col_idx
            elif 'transmission' in val or 'status' in val:
                header_map['transmission_status'] = col_idx
            elif 'available' in val or 'expected' in val:
                header_map['expected_available'] = col_idx
            elif 'capacity' in val and 'mw' in val:
                if 'hosting_capacity' not in header_map:
                    header_map['hosting_capacity'] = col_idx

        if 'name' not in header_map:
            print(f"    No REZ name column found in {sheet_name}")
            continue

        # Read data rows
        for row_idx in range(header_row + 1, ws.max_row + 1):
            name_cell = ws.cell(row=row_idx, column=header_map['name']).value
            if not name_cell:
                continue

            rez_id = normalize_rez_name(name_cell)
            if not rez_id:
                skipped += 1
                continue

            hosting = None
            if 'hosting_capacity' in header_map:
                v = ws.cell(row=row_idx, column=header_map['hosting_capacity']).value
                if v is not None:
                    try:
                        hosting = float(v)
                    except (ValueError, TypeError):
                        pass

            connection = None
            if 'connection_capacity' in header_map:
                v = ws.cell(row=row_idx, column=header_map['connection_capacity']).value
                if v is not None:
                    try:
                        connection = float(v)
                    except (ValueError, TypeError):
                        pass

            transmission = None
            if 'transmission_status' in header_map:
                transmission = str(ws.cell(row=row_idx, column=header_map['transmission_status']).value or '')

            expected = None
            if 'expected_available' in header_map:
                expected = str(ws.cell(row=row_idx, column=header_map['expected_available']).value or '')

            # Upsert
            conn.execute("""
                INSERT INTO rez_isp_data (rez_id, isp_year, hosting_capacity_mw, connection_capacity_mw, transmission_status, expected_available)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(rez_id, isp_year) DO UPDATE SET
                    hosting_capacity_mw = excluded.hosting_capacity_mw,
                    connection_capacity_mw = excluded.connection_capacity_mw,
                    transmission_status = excluded.transmission_status,
                    expected_available = excluded.expected_available,
                    import_date = date('now')
            """, (rez_id, isp_year, hosting, connection, transmission, expected))
            imported += 1

    conn.commit()
    wb.close()
    print(f"  Imported: {imported} REZ records, Skipped: {skipped}")
    return imported


def main():
    parser = argparse.ArgumentParser(description='Import AEMO ISP data')
    parser.add_argument('--file', type=str, help='Path to specific ISP Excel file')
    parser.add_argument('--year', type=int, default=2024, help='ISP publication year (default: 2024)')
    args = parser.parse_args()

    conn = get_connection()
    ensure_isp_table(conn)

    started = datetime.now().isoformat()

    if args.file:
        if not os.path.exists(args.file):
            print(f"ERROR: File not found: {args.file}")
            sys.exit(1)
        count = import_isp_excel(conn, args.file, args.year)
    else:
        # Scan downloads directory for ISP files
        count = 0
        if os.path.exists(DOWNLOADS_DIR):
            for fname in sorted(os.listdir(DOWNLOADS_DIR)):
                if fname.lower().endswith('.xlsx') and 'isp' in fname.lower():
                    fpath = os.path.join(DOWNLOADS_DIR, fname)
                    count += import_isp_excel(conn, fpath, args.year)
        if count == 0:
            print(f"No ISP Excel files found in {DOWNLOADS_DIR}")
            print("Place ISP appendix files there or use --file flag.")

    # Log the run
    completed = datetime.now().isoformat()
    try:
        conn.execute("""
            INSERT INTO import_runs (source, started_at, completed_at, status, records_imported)
            VALUES (?, ?, ?, ?, ?)
        """, ('aemo_isp', started, completed, 'completed', count))
        conn.commit()
    except Exception:
        pass

    conn.close()
    print(f"\nTotal ISP records imported: {count}")


if __name__ == '__main__':
    main()
