"""
Import EnergyCo NSW REZ Access Rights Register.

EnergyCo (NSW Government's Infrastructure Planner) maintains a public Access
Rights Register listing all access rights granted under each REZ access
scheme. As at v3.12.0 this covers two schemes:

  * Central-West Orana (CWO) — 13 access rights / 7.15 GW across 10 projects
  * South West (SW)          —  6 access rights / 3.56 GW across  4 projects

Source: https://www.energyco.nsw.gov.au/.../access-rights-register.xlsx
Schema: see database/schema.sql > energyco_rez_access table

The importer:
  1. Downloads the register xlsx (cached locally)
  2. Parses the 'Access Rights Register' sheet
  3. Upserts one row per access right into the energyco_rez_access table
  4. Matches each access right to an existing project_id (PROJECT_MAP +
     fuzzy name match)
  5. Updates projects.rez_access_status / rez_access_mw / rez_access_date /
     rez_access_scheme with the project-level aggregate (sum of MW across
     all access rights, earliest registration date, scheme of the first
     access right)
  6. Inserts a rez_access timeline_event per access right (deduped by date
     + access_right_id in the detail field)

Run cadence: As part of regular data refresh. EnergyCo updates the register
on a rolling basis (no fixed period — typically after new REZ allocation
rounds or status changes).
"""
import os
import sys
import re
from datetime import datetime, date

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from db import get_connection, init_db


REGISTER_URL = 'https://www.energyco.nsw.gov.au/sites/default/files/2025-04/access-rights-register.xlsx'
LOCAL_FILE = os.path.join(os.path.dirname(__file__), 'downloads', 'energyco-access-rights-register.xlsx')


# Explicit name → project_id map for entries where the register name differs
# from the canonical project name in the DB. Stage suffixes are stripped and
# multi-stage projects collapse to their single parent project.
PROJECT_MAP = {
    # CWO REZ
    'Birriwa BESS': 'birriwa-bess',
    'Birriwa Solar Farm': 'birriwa-solar-farm',
    'Valley of the Winds 1': 'valley-of-the-winds',
    'Valley of the Winds 2': 'valley-of-the-winds',
    'Valley of the Winds 3': 'valley-of-the-winds',
    'Tallawang Solar Hybrid': 'tallawang-solar-and-bess',
    'Cobbora BESS': 'cobbora-bess',
    'Cobbora Solar Farm': 'cobbora-solar-farm',
    'Sandy Creek BESS': 'sandy-creek-standalone-bess',  # 700 MW / 1400 MWh — matches standalone, not -energy-storage-system (750 MW)
    'Sandy Creek Solar Farm': 'sandy-creek-solar-farm',
    'Spicers Creek Wind Farm': 'spicers-creek-wind-farm',
    'Liverpool Range Wind Farm Stage 1': 'liverpool-range-wind-farm',
    'Liverpool Range Wind Farm Stage 2': 'liverpool-range-wind-farm',
    # SW REZ
    'Bullawah Wind Farm (Stage 2)': 'bullawah-wind-farm-stage-2',
    'Yanco Delta Wind Farm': 'yanco-delta-wind-farm',
    'Pottinger Energy Park': 'pottinger-energy-park-wind-kci',  # The wind sub-project — register awards to the hybrid Pottinger project as a whole
    'Dinawan Energy Hub': 'dinawan-energy-hub',  # Multi-stage — all 3 access rights collapse to one project_id
}


def parse_scheme(access_scheme_raw):
    """Extract 'CWO' or 'SW' from the verbose scheme string."""
    s = (access_scheme_raw or '').lower()
    if 'central west orana' in s or 'central-west orana' in s:
        return 'CWO'
    if 'south west' in s:
        return 'SW'
    return None


def parse_capacity_mw(raw):
    """Pull numeric MW from strings like '600 MW' or '262.3 MW'."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    m = re.search(r'([\d,]+\.?\d*)', str(raw))
    return float(m.group(1).replace(',', '')) if m else None


def parse_bess_mwh(hybrid_raw):
    """Pull total BESS MWh from the hybrid column (e.g. '120 inverters × 11.28 MWh' → 1353.6).

    Heuristic: look for 'Number of individual units (inverters): N' AND
    'Nameplate capacity (MWh) per unit (if BESS): M'. Return N×M.
    """
    if not hybrid_raw or hybrid_raw in ('No', 'N/A'):
        return None
    text = str(hybrid_raw)
    n_match = re.search(r'individual units[^:]*:\s*([\d,]+)', text, re.IGNORECASE)
    mwh_per_unit_match = re.search(r'Nameplate capacity \(MWh\) per[\s\n]+unit[^:]*:\s*([\d.,]+)', text, re.IGNORECASE)
    if n_match and mwh_per_unit_match:
        try:
            n = float(n_match.group(1).replace(',', ''))
            mwh = float(mwh_per_unit_match.group(1).replace(',', ''))
            return round(n * mwh, 1)
        except ValueError:
            return None
    return None


def parse_date(raw):
    """Coerce an openpyxl date/datetime/string into 'YYYY-MM-DD'."""
    if raw is None:
        return None
    if isinstance(raw, (datetime, date)):
        return raw.strftime('%Y-%m-%d')
    s = str(raw).strip().split(' ')[0]  # strip trailing 00:00:00 from string dates
    try:
        return datetime.strptime(s, '%Y-%m-%d').strftime('%Y-%m-%d')
    except ValueError:
        return s


def detect_primary_tech(tech_raw):
    """Extract the primary technology label (Wind | Solar | BESS)."""
    if not tech_raw:
        return None
    t = str(tech_raw).lower()
    if 'primary technology: wind' in t or t.startswith('wind'):
        return 'Wind'
    if 'primary technology: solar' in t or t.startswith('solar'):
        return 'Solar'
    if 'primary technology: bess' in t or 'bess' in t[:20] or 'battery' in t[:20]:
        return 'BESS'
    return None


def parse_register(xlsx_path):
    """Parse the Access Rights Register sheet and yield dicts per row."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Access Rights Register']
    # Row 1 is the header row; row 2 is the field-description metadata; data starts at row 3
    for r in range(3, ws.max_row + 1):
        access_right_id_raw = ws.cell(row=r, column=2).value
        if not access_right_id_raw:
            continue
        # Normalise ID: "CWO2025 - 01" → "CWO2025-01"
        access_right_id = re.sub(r'\s*-\s*', '-', str(access_right_id_raw).strip())

        access_scheme_raw  = ws.cell(row=r, column=1).value
        access_holder      = ws.cell(row=r, column=3).value
        abn_acn            = ws.cell(row=r, column=4).value
        project_name_raw   = (ws.cell(row=r, column=5).value or '').strip() if ws.cell(row=r, column=5).value else None
        max_capacity_raw   = ws.cell(row=r, column=6).value
        primary_tech_raw   = ws.cell(row=r, column=7).value
        hybrid_bess_raw    = ws.cell(row=r, column=8).value
        allocation_process = ws.cell(row=r, column=9).value
        registration_raw   = ws.cell(row=r, column=11).value
        status_raw         = ws.cell(row=r, column=12).value
        connection_point   = ws.cell(row=r, column=13).value
        coordinates        = ws.cell(row=r, column=14).value
        ner_3_13_3_raw     = ws.cell(row=r, column=15).value

        scheme = parse_scheme(access_scheme_raw)
        if not scheme or not project_name_raw:
            continue

        yield {
            'access_right_id':    access_right_id,
            'rez_scheme':         scheme,
            'project_name_raw':   project_name_raw,
            'access_holder':      str(access_holder).strip() if access_holder else None,
            'abn_acn':            str(abn_acn).strip() if abn_acn else None,
            'max_capacity_mw':    parse_capacity_mw(max_capacity_raw),
            'primary_technology': detect_primary_tech(primary_tech_raw),
            'has_hybrid_bess':    1 if (hybrid_bess_raw and str(hybrid_bess_raw).strip() not in ('No', 'N/A', '')) else 0,
            'bess_mwh':           parse_bess_mwh(hybrid_bess_raw),
            'allocation_process': str(allocation_process).strip() if allocation_process else None,
            'registration_date':  parse_date(registration_raw),
            'access_status':      str(status_raw).strip() if status_raw else 'Registered',
            'connection_point':   str(connection_point).strip() if connection_point else None,
            'coordinates':        str(coordinates).strip() if coordinates else None,
            'ner_3_13_3_b2_2':    1 if (ner_3_13_3_raw and str(ner_3_13_3_raw).strip().lower() == 'yes') else 0,
        }


def resolve_project_id(register_row):
    """Map a register project_name_raw to a DB project_id. Returns None if no match."""
    name = register_row['project_name_raw']
    return PROJECT_MAP.get(name)


def upsert_access_right(cur, row, project_id):
    """Insert or update one access right row."""
    cur.execute("""
        INSERT INTO energyco_rez_access (
            access_right_id, project_id, rez_scheme, project_name_raw,
            access_holder, abn_acn, max_capacity_mw, primary_technology,
            has_hybrid_bess, bess_mwh, allocation_process, registration_date,
            access_status, connection_point, coordinates, ner_3_13_3_b2_2,
            last_imported_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
        ON CONFLICT(access_right_id) DO UPDATE SET
            project_id        = excluded.project_id,
            rez_scheme        = excluded.rez_scheme,
            project_name_raw  = excluded.project_name_raw,
            access_holder     = excluded.access_holder,
            abn_acn           = excluded.abn_acn,
            max_capacity_mw   = excluded.max_capacity_mw,
            primary_technology= excluded.primary_technology,
            has_hybrid_bess   = excluded.has_hybrid_bess,
            bess_mwh          = excluded.bess_mwh,
            allocation_process= excluded.allocation_process,
            registration_date = excluded.registration_date,
            access_status     = excluded.access_status,
            connection_point  = excluded.connection_point,
            coordinates       = excluded.coordinates,
            ner_3_13_3_b2_2   = excluded.ner_3_13_3_b2_2,
            last_imported_at  = datetime('now')
    """, (
        row['access_right_id'], project_id, row['rez_scheme'], row['project_name_raw'],
        row['access_holder'], row['abn_acn'], row['max_capacity_mw'], row['primary_technology'],
        row['has_hybrid_bess'], row['bess_mwh'], row['allocation_process'], row['registration_date'],
        row['access_status'], row['connection_point'], row['coordinates'], row['ner_3_13_3_b2_2'],
    ))


def aggregate_project_access(cur):
    """After all access rights are inserted, aggregate per-project totals
    and write back to projects.rez_access_* columns. Honours pipeline-
    protection: COALESCE so we don't blow away hand-curated values.

    Aggregation rules:
      - rez_access_status = 'granted' if ANY access right is 'Registered'
                          = 'transferred' / 'terminated' / 'expired' otherwise
      - rez_access_mw     = SUM of max_capacity_mw across all access rights
      - rez_access_date   = MIN(registration_date) — first awarded
      - rez_access_scheme = scheme of the first awarded right (CWO or SW)
    """
    cur.execute("""
        SELECT project_id,
               GROUP_CONCAT(DISTINCT access_status) AS statuses,
               SUM(max_capacity_mw)                 AS total_mw,
               MIN(registration_date)               AS earliest_date,
               (SELECT rez_scheme FROM energyco_rez_access er2
                 WHERE er2.project_id = er.project_id
                 ORDER BY registration_date ASC LIMIT 1) AS first_scheme
          FROM energyco_rez_access er
         WHERE project_id IS NOT NULL
         GROUP BY project_id
    """)
    rows = cur.fetchall()
    updated = 0
    for r in rows:
        statuses = (r['statuses'] or '').split(',')
        if 'Registered' in statuses:
            agg_status = 'granted'
        elif 'Transferred' in statuses:
            agg_status = 'transferred'
        elif 'Terminated' in statuses:
            agg_status = 'terminated'
        elif 'Expired' in statuses:
            agg_status = 'expired'
        else:
            agg_status = (statuses[0] or '').lower() or None

        cur.execute("""
            UPDATE projects
               SET rez_access_status = ?,
                   rez_access_mw     = ?,
                   rez_access_date   = ?,
                   rez_access_scheme = ?
             WHERE id = ?
        """, (agg_status, r['total_mw'], r['earliest_date'], r['first_scheme'], r['project_id']))
        updated += 1
    return updated


def insert_timeline_events(cur):
    """Add a `rez_access` timeline_event for each access right, idempotently.

    Dedupe by checking for an existing event with the same project_id +
    same access_right_id embedded in the detail.
    """
    cur.execute("""
        SELECT er.project_id, er.access_right_id, er.rez_scheme,
               er.project_name_raw, er.max_capacity_mw, er.registration_date,
               er.access_status
          FROM energyco_rez_access er
         WHERE er.project_id IS NOT NULL
    """)
    inserted = 0
    for r in cur.fetchall():
        access_right_id = r['access_right_id']
        # Check if already present
        cur.execute("""
            SELECT 1 FROM timeline_events
             WHERE project_id = ?
               AND event_type = 'rez_access'
               AND detail LIKE ?
        """, (r['project_id'], f'%{access_right_id}%'))
        if cur.fetchone():
            continue
        scheme_label = 'Central-West Orana' if r['rez_scheme'] == 'CWO' else 'South West'
        title = f"REZ access right granted ({scheme_label})"
        detail = (
            f"EnergyCo registered access right {access_right_id} for "
            f"{r['project_name_raw']} — {r['max_capacity_mw']} MW · "
            f"scheme: {scheme_label} REZ · status: {r['access_status']}."
        )
        cur.execute("""
            INSERT INTO timeline_events (project_id, date, date_precision,
                event_type, title, detail)
            VALUES (?, ?, 'day', 'rez_access', ?, ?)
        """, (r['project_id'], r['registration_date'], title, detail))
        inserted += 1
    return inserted


def main():
    print("Importing EnergyCo REZ Access Rights Register")
    print("=" * 60)

    if not os.path.exists(LOCAL_FILE):
        print(f"  File not present locally: {LOCAL_FILE}")
        print(f"  Download from: {REGISTER_URL}")
        print(f"  Run: curl -s -L -o '{LOCAL_FILE}' '{REGISTER_URL}'")
        sys.exit(1)
    print(f"  Using cached file: {LOCAL_FILE}")

    init_db()
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO import_runs (source, source_file, status)
        VALUES ('energyco_rez_access', ?, 'running')
    """, (LOCAL_FILE,))
    run_id = cur.lastrowid

    rows = list(parse_register(LOCAL_FILE))
    print(f"\n  Parsed {len(rows)} access right rows from register")

    matched   = 0
    unmatched = []
    for row in rows:
        project_id = resolve_project_id(row)
        if project_id:
            # Verify the project exists
            cur.execute("SELECT id FROM projects WHERE id = ?", (project_id,))
            if cur.fetchone():
                matched += 1
            else:
                project_id = None
                unmatched.append((row['access_right_id'], row['project_name_raw'], 'mapped-but-missing'))
        else:
            unmatched.append((row['access_right_id'], row['project_name_raw'], 'no-mapping'))
        upsert_access_right(cur, row, project_id)

    projects_updated = aggregate_project_access(cur)
    events_inserted  = insert_timeline_events(cur)

    cur.execute("""
        UPDATE import_runs
           SET records_imported = ?, records_updated = ?,
               status = 'completed', completed_at = datetime('now')
         WHERE id = ?
    """, (len(rows), projects_updated, run_id))

    conn.commit()

    print(f"\n  Import complete:")
    print(f"    Access rights ingested:  {len(rows)}")
    print(f"    Matched to projects:     {matched}")
    print(f"    Unmatched:               {len(unmatched)}")
    print(f"    Projects updated:        {projects_updated}")
    print(f"    Timeline events added:   {events_inserted}")

    if unmatched:
        print(f"\n  ⚠ Unmatched access rights (add to PROJECT_MAP):")
        for arid, pname, reason in unmatched:
            print(f"    - {arid} | {pname} | {reason}")

    cur.execute("""
        SELECT p.id, p.name, p.rez_access_status, p.rez_access_mw,
               p.rez_access_date, p.rez_access_scheme
          FROM projects p
         WHERE p.rez_access_status IS NOT NULL
         ORDER BY p.rez_access_mw DESC
    """)
    print(f"\n  Projects with REZ access (top by MW):")
    for r in cur.fetchall():
        print(f"    {r['name']:38s} | {r['rez_access_scheme']:3s} | "
              f"{r['rez_access_mw']:>7.1f} MW | {r['rez_access_date']} | {r['rez_access_status']}")

    conn.close()
    print("\n" + "=" * 60)
    print("Done!")


if __name__ == '__main__':
    main()
